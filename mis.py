from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter

import os


# =========================================================
# MIS HEADERS
# =========================================================

MIS_HEADERS = [
    "Request Number",
    "Request Date",
    "Requester Name",
    "Requester Email",
    "Employee ID",
    "Department",
    "Designation",
    "Requirement Description",
    "Business Requirement",
    "BOQ",
    "Approver Emails",
    "Status",
    "Attachment Filename",
    "Created At",
    "Updated At"
]


# =========================================================
# COLUMN WIDTHS
# =========================================================

COLUMN_WIDTHS = [
    24,  # Request Number
    15,  # Request Date
    22,  # Requester Name
    30,  # Requester Email
    15,  # Employee ID
    18,  # Department
    18,  # Designation
    45,  # Requirement Description
    45,  # Business Requirement
    35,  # BOQ
    35,  # Approver Emails
    15,  # Status
    35,  # Attachments
    22,  # Created At
    22   # Updated At
]


# =========================================================
# CREATE / FORMAT MIS WORKSHEET
# =========================================================

def create_mis_workbook(mis_file):
    """
    Create a new MIS Excel workbook with headers and formatting.
    """

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Procurement MIS"

    # -----------------------------------------------------
    # Headers
    # -----------------------------------------------------

    worksheet.append(MIS_HEADERS)

    # -----------------------------------------------------
    # Header styling
    # -----------------------------------------------------

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    thin_border = Border(
        bottom=Side(
            style="thin",
            color="D9E2F3"
        )
    )

    for cell in worksheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border

    # -----------------------------------------------------
    # Freeze header
    # -----------------------------------------------------

    worksheet.freeze_panes = "A2"

    # -----------------------------------------------------
    # Column widths
    # -----------------------------------------------------

    for index, width in enumerate(
        COLUMN_WIDTHS,
        start=1
    ):

        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = width

    # -----------------------------------------------------
    # Header row height
    # -----------------------------------------------------

    worksheet.row_dimensions[1].height = 30

    # -----------------------------------------------------
    # Save
    # -----------------------------------------------------

    workbook.save(
        mis_file
    )


# =========================================================
# FIND REQUEST ROW
# =========================================================

def find_request_row(worksheet, request_number):
    """
    Find an existing MIS row using Request Number.

    Request Number is stored in column A.
    """

    if not request_number:
        return None

    request_number = str(
        request_number
    ).strip()

    # Start from row 2 because row 1 is header.
    for row in range(
        2,
        worksheet.max_row + 1
    ):

        existing_request_number = (
            worksheet.cell(
                row=row,
                column=1
            ).value
        )

        if existing_request_number is None:
            continue

        if str(
            existing_request_number
        ).strip() == request_number:

            return row

    return None


# =========================================================
# REQUEST DATA TO ROW
# =========================================================

def request_data_to_row(request_data):
    """
    Convert request dictionary into the exact MIS row format.
    """

    return [

        request_data.get(
            "request_number"
        ),

        request_data.get(
            "request_date"
        ),

        request_data.get(
            "requester_name"
        ),

        request_data.get(
            "requester_email"
        ),

        request_data.get(
            "employee_id"
        ),

        request_data.get(
            "department"
        ),

        request_data.get(
            "designation"
        ),

        request_data.get(
            "item_description"
        ),

        request_data.get(
            "business_requirement"
        ),

        request_data.get(
            "boq"
        ),

        request_data.get(
            "approver_emails"
        ),

        request_data.get(
            "status"
        ),

        request_data.get(
            "attachment_filename"
        ),

        request_data.get(
            "created_at"
        ),

        request_data.get(
            "updated_at"
        )
    ]


# =========================================================
# STYLE DATA ROW
# =========================================================

def style_data_row(worksheet, row_number):
    """
    Apply formatting to a MIS data row.
    """

    for column in range(
        1,
        len(MIS_HEADERS) + 1
    ):

        cell = worksheet.cell(
            row=row_number,
            column=column
        )

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

    # -----------------------------------------------------
    # Date formatting
    # -----------------------------------------------------

    # Request Date
    worksheet.cell(
        row=row_number,
        column=2
    ).number_format = "yyyy-mm-dd"

    # Created At
    worksheet.cell(
        row=row_number,
        column=14
    ).number_format = "yyyy-mm-dd hh:mm:ss"

    # Updated At
    worksheet.cell(
        row=row_number,
        column=15
    ).number_format = "yyyy-mm-dd hh:mm:ss"


# =========================================================
# SAVE / UPDATE MIS
# =========================================================

def save_request_to_mis(request_data, mis_file):
    """
    Add or update a procurement request in the Excel MIS.

    PostgreSQL remains the master database.

    Behaviour:

        If Request Number does NOT exist:
            Create a new row.

        If Request Number already exists:
            Update the existing row.

    This prevents duplicate MIS rows when a request moves
    between statuses such as:

        SUBMITTED
        QUERY
        APPROVED
        RETURNED
    """

    try:

        # =================================================
        # VALIDATE REQUEST DATA
        # =================================================

        if not request_data:

            raise ValueError(
                "request_data is empty."
            )

        request_number = (
            request_data.get(
                "request_number"
            )
        )

        if not request_number:

            raise ValueError(
                "Request Number is required for MIS."
            )

        request_number = str(
            request_number
        ).strip()

        # =================================================
        # CREATE MIS FOLDER
        # =================================================

        mis_folder = os.path.dirname(
            mis_file
        )

        if mis_folder:

            os.makedirs(
                mis_folder,
                exist_ok=True
            )

        # =================================================
        # CREATE WORKBOOK IF IT DOES NOT EXIST
        # =================================================

        if not os.path.exists(
            mis_file
        ):

            create_mis_workbook(
                mis_file
            )

        # =================================================
        # OPEN WORKBOOK
        # =================================================

        workbook = load_workbook(
            mis_file
        )

        # =================================================
        # GET / CREATE WORKSHEET
        # =================================================

        if "Procurement MIS" in workbook.sheetnames:

            worksheet = workbook[
                "Procurement MIS"
            ]

        else:

            worksheet = workbook.create_sheet(
                "Procurement MIS"
            )

            worksheet.append(
                MIS_HEADERS
            )

            # Header styling
            header_fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            header_font = Font(
                color="FFFFFF",
                bold=True
            )

            thin_border = Border(
                bottom=Side(
                    style="thin",
                    color="D9E2F3"
                )
            )

            for cell in worksheet[1]:

                cell.fill = header_fill

                cell.font = header_font

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                cell.border = thin_border

            worksheet.freeze_panes = "A2"

            for index, width in enumerate(
                COLUMN_WIDTHS,
                start=1
            ):

                worksheet.column_dimensions[
                    get_column_letter(index)
                ].width = width

            worksheet.row_dimensions[1].height = 30

        # =================================================
        # CONVERT REQUEST DATA TO ROW
        # =================================================

        row_data = request_data_to_row(
            request_data
        )

        # =================================================
        # FIND EXISTING REQUEST
        # =================================================

        existing_row = find_request_row(
            worksheet,
            request_number
        )

        # =================================================
        # UPDATE EXISTING REQUEST
        # =================================================

        if existing_row:

            for column_index, value in enumerate(
                row_data,
                start=1
            ):

                worksheet.cell(
                    row=existing_row,
                    column=column_index
                ).value = value

            style_data_row(
                worksheet,
                existing_row
            )

            print(
                f"MIS UPDATED: "
                f"{request_number} "
                f"-> status={request_data.get('status')}"
            )

        # =================================================
        # ADD NEW REQUEST
        # =================================================

        else:

            worksheet.append(
                row_data
            )

            new_row = worksheet.max_row

            style_data_row(
                worksheet,
                new_row
            )

            print(
                f"MIS INSERTED: "
                f"{request_number} "
                f"-> status={request_data.get('status')}"
            )

        # =================================================
        # UPDATE AUTO FILTER
        # =================================================

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        # =================================================
        # SAVE WORKBOOK
        # =================================================

        workbook.save(
            mis_file
        )

        # =================================================
        # CLOSE WORKBOOK
        # =================================================

        workbook.close()

        print(
            f"MIS SAVE SUCCESS: {request_number}"
        )

        return True

    except Exception as e:

        print(
            "MIS ERROR:",
            repr(e)
        )

        return False